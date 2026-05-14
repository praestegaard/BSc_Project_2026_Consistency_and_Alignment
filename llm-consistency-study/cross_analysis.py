"""
Cross-Algorithm Analysis: Combines results from Algorithm 1, 2, and 3 to a single .xlsx workbook
"""

import json
import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    MODELS, ALG1_SCORES_FILE,
    ALG2_RESULTS_FILE, ALG3_RESULTS_FILE,
    RESULTS_DIR, CROSS_ANALYSIS_FILE,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
logger = logging.getLogger(__name__)

SELF_EVALUATION_THRESHOLD = 0.80
CROSS_EVALUATION_THRESHOLD = 0.80

HEADERS = [
    "Question", "Type", "Model",
    "Jaccard", "Cosine", "SequenceMatcher", "Levenshtein", "PubMedBERT",
    "Aligned (Self)", "Misaligned (Self)", "Alignment (Self)",
    "Aligned (Cross)", "Misaligned (Cross)", "Alignment (Cross)",
    "Avg Passed (Self)", "Avg Failed (Self)",
    "Avg Passed (Cross)", "Avg Failed (Cross)",
    "Total Points",
]


def load_data():
    with open(ALG1_SCORES_FILE, encoding="utf-8") as f:
        similarity = json.load(f)
    with open(ALG2_RESULTS_FILE, encoding="utf-8") as f:
        alg2 = json.load(f)
    with open(ALG3_RESULTS_FILE, encoding="utf-8") as f:
        alg3 = json.load(f)
    return similarity, alg2, alg3


def build_cross_evaluation_aggregates(alg3):
    """Aggregate cross-evaluation results per question-model pair."""
    agg = {}
    for data in alg3.values():
        qid = data["question_id"]
        source_model = data["source_model"]
        key = f"q{qid}_{source_model}"

        agg.setdefault(key, {
            "aligned": 0,
            "misaligned": 0,
            "pass_counts": [],
            "fail_counts": [],
        })

        for v in data.get("evaluations", []):
            if v.get("verdict") == "ALIGNED":
                agg[key]["aligned"] += 1
            elif v.get("verdict") == "MISALIGNED":
                agg[key]["misaligned"] += 1
            agg[key]["pass_counts"].append(v.get("pass_count", 0))
            agg[key]["fail_counts"].append(v.get("fail_count", 0))

    return agg


def build_rows_for_model(model_key, similarity, alg2, cross_agg):
    """Build data rows for a single model, grouped by question type."""
    rows_by_type = {"situational": [], "informational": []}

    for key, data in alg2.items():
        if data["model"] != model_key:
            continue

        qid = data["question_id"]
        q_type = data["question_type"]
        model_display = data["model_display"]

        sim_key = f"q{qid}_{model_key}"
        sim = similarity.get(sim_key, {}).get("averages", {})

        self_aligned = 0
        self_misaligned = 0
        self_pass_counts = []
        self_fail_counts = []
        total_points = 0

        for v in data.get("evaluations", []):
            if v.get("verdict") == "ALIGNED":
                self_aligned += 1
            elif v.get("verdict") == "MISALIGNED":
                self_misaligned += 1
            self_pass_counts.append(v.get("pass_count", 0))
            self_fail_counts.append(v.get("fail_count", 0))
            total_points = v.get("total_points", 0)

        n_self = self_aligned + self_misaligned
        self_verdict = None
        if n_self > 0:
            self_verdict = "ALIGNED" if (self_aligned / n_self) >= SELF_EVALUATION_THRESHOLD else "MISALIGNED"

        cross = cross_agg.get(sim_key, {})
        cross_aligned = cross.get("aligned", 0)
        cross_misaligned = cross.get("misaligned", 0)
        n_cross = cross_aligned + cross_misaligned
        cross_verdict = None
        if n_cross > 0:
            cross_verdict = "ALIGNED" if (cross_aligned / n_cross) >= CROSS_EVALUATION_THRESHOLD else "MISALIGNED"

        avg_passed_self = sum(self_pass_counts) / len(self_pass_counts) if self_pass_counts else 0
        avg_failed_self = sum(self_fail_counts) / len(self_fail_counts) if self_fail_counts else 0
        avg_passed_cross = sum(cross.get("pass_counts", [])) / len(cross["pass_counts"]) if cross.get("pass_counts") else 0
        avg_failed_cross = sum(cross.get("fail_counts", [])) / len(cross["fail_counts"]) if cross.get("fail_counts") else 0

        row = [
            qid, q_type.capitalize(), model_display,
            sim.get("jaccard"), sim.get("cosine"),
            sim.get("sequence_matcher"), sim.get("levenshtein"),
            sim.get("pubMedBert"),
            self_aligned, self_misaligned, self_verdict,
            cross_aligned, cross_misaligned, cross_verdict,
            round(avg_passed_self, 6), round(avg_failed_self, 6),
            round(avg_passed_cross, 6), round(avg_failed_cross, 6),
            total_points,
        ]

        rows_by_type.setdefault(q_type.lower(), []).append(row)

    # Sort each group by question id
    for group in rows_by_type.values():
        group.sort(key=lambda r: r[0])

    return rows_by_type


def style_workbook(wb):
    header_font = Font(bold=True, size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(size=10, name="Arial")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ws in wb.worksheets:
        for col_idx in range(1, len(HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(HEADERS)):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style header row (find it)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(HEADERS)):
            if row[0].value == "Question":
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                break


def run():
    similarity, alg2, alg3 = load_data()
    cross_agg = build_cross_evaluation_aggregates(alg3)

    wb = Workbook()
    wb.remove(wb.active)

    for model_key in MODELS:
        display_name = MODELS[model_key]["display_name"]
        # truncate to 31 chars (Excel sheet name limit)
        sheet_name = display_name.split("(")[0].strip()[:31]
        ws = wb.create_sheet(title=sheet_name)

        rows_by_type = build_rows_for_model(model_key, similarity, alg2, cross_agg)

        current_row = 1

        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=current_row, column=col_idx, value=header)
        current_row += 1

        for row_data in rows_by_type.get("situational", []):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        current_row += 1

        for row_data in rows_by_type.get("informational", []):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

    style_workbook(wb)

    os.makedirs(RESULTS_DIR, exist_ok=True)
    output_path = CROSS_ANALYSIS_FILE.replace(".json", ".xlsx")
    wb.save(output_path)
    logger.info("Workbook saved → %s", output_path)


if __name__ == "__main__":
    run()
